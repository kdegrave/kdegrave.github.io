from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

# Define cell hex colors to use
yellow = 'fefe96'
orange = 'ffdfba'
green = 'a9ebc3'
blue = 'bae1ff'
red = 'ffb3ba'

# Define the alignment of header text
alignment = Alignment(horizontal='center', vertical='center')

# Define the header cell borders
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

def formatcells(ws, crange, value, border, fill, alignment):

    ws.merge_cells(start_row=crange[0], start_column=crange[1],
                       end_row=crange[2], end_column=crange[3])

    cell = ws.cell(row=crange[0], column=crange[1])

    if value:
        cell = ws.cell(row=crange[0], column=crange[1], value=value)

    if alignment:
        cell.alignment = alignment

    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')

    if border:
        for i in range(crange[0], crange[2] + 1):
            for j in range(crange[1], crange[3] + 1):
                cell = ws.cell(row=i, column=j)
                cell.border = border

def formatlines(ws, crange):
    if border:
        for i in range(crange[0], crange[2] + 1):
            for j in range(crange[1], crange[3] + 1):
                cell = ws.cell(row=i, column=j)
                cell.border = Border(bottom=Side(style='medium'))

def formatcolumns(ws):
    dims = {}
    min_width = 12
    for row in ws.rows:
        for cell in row:
            try:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
            except: pass

    for col, value in dims.items():
        if value < min_width:
            value = min_width
        ws.column_dimensions[col].width = 1.0*(value)
    return ws

def formatdecimals(ws):
    for row in ws.rows:
        for cell in row:
            try:
                if cell.value % 1 > 0:
                    cell.number_format = '0.00'
                else: cell.number_format = '#,###'
            except: pass
    return ws

def formatpercentages(ws):
    for row in ws.rows:
        for cell in row:
            try:
                if '%%' in cell.value:
                    cell.value = cell.value.strip('%%')
                    cell.value = float(cell.value)/100
                    if cell.value < 0:
                        cell.fill = PatternFill(start_color='FFC7CE',
                                end_color='FFC7CE', fill_type='solid')
                    if cell.value % 1 > 0:
                        cell.number_format = '0.00%'
                    else: cell.number_format = '#,###%'
            except: pass
    return ws